"""
© BRAIN STATION 23 | Design and Development: Md. Ariful Islam (BS1121)
"""
import os
import openpyxl
from openpyxl import Workbook
from testdata.abstract.test_data import TestData
from constants.common_constants import CommonConstants

class XLtestData(TestData):
    """
    This XLtestData class is responsible for reading data from xlsx files, 
    arrange those data for Data Driven Testing. This class also have functions
    to convert the XL Data Set into list values and convert the Data Set header 
    as list variable names.  
    © BRAIN STATION 23 | Design and Development: Md. Ariful Islam (BS1121)
    """
    workbook_path = CommonConstants.WORKBOOK_FILE_PATH
    workbook_file = openpyxl.load_workbook(workbook_path)
    sheet_file: Workbook = None
    xl_data_dictionary: dict = {}

    def display_all_data(self):
        self.__select_sheet()
        rows = self.sheet_file.max_row
        columns = self.sheet_file.max_column
        for curr_row in range(1, rows + 1):
            for curr_column in range(1, columns + 1):
                print(self.sheet_file.cell(row=curr_row, column=curr_column).value, end='     ')
            print('\n')
        print('------END------')

    def arrange_data_set(self, arrange_type: str):
        if arrange_type == 'vertical':
            return self.__arrange_data_by_variable_value_list()
        elif arrange_type == 'horizontal':
            pass
        else:
            return 'Type unknown!'

    def variable_validator(self, first_row_value: str):
        return super().variable_validator(first_row_value)

    def variable_converter(self, first_row_value: str):
        return super().variable_converter(first_row_value)

    def get_all_data(self):
        return super().get_all_data()

    def get_data_by_column(self, column_name: str):
        return super().get_data_by_column(column_name)

    def get_data_with_condition(self, column_name: str, condition: str, operator: str):
        return super().get_data_with_condition(column_name, condition, operator)

    def __select_sheet(self, sheet_name: str=None):
        sheet_names = self.workbook_file.sheetnames
        print('These are the available sheet names', sheet_names)
        if sheet_name is None:
            sheet_name = input('Provide your sheet name here and press ENTER:')
        self.sheet_file = self.workbook_file[sheet_name]
        print('-------Sheet Found!-------')

    def __arrange_data_by_variable_value_list(self):
        self.xl_data_dictionary
        self.__select_sheet()
        rows = self.sheet_file.max_row
        columns = self.sheet_file.max_column
        curr_variable: str = None
        for curr_column in range(1, columns + 1):
            for curr_row in range(1, rows + 1):
                current_cell_value = str(self.sheet_file.cell(row=curr_row, column=curr_column).value)
                '\n                If it is the first Row of a sheet, each cell value of that \n                Row will be identified as a variable. \n                '
                if curr_row == 1:
                    curr_variable = self.variable_converter(current_cell_value)
                    if self.variable_validator(curr_variable):
                        self.xl_data_dictionary[curr_variable] = []
                    else:
                        print(f'\n                                ***********************************************************\n                                    !!!Given name for variable: {curr_variable}. !!!\n                                This element is not suitable to be decleared as a Variable!\n                                Please, try to follow the standards!\n                                ***********************************************************\n                                ')
                        break
                elif curr_variable is not None:
                    self.xl_data_dictionary[curr_variable].append(current_cell_value)
                else:
                    return 'Something went wrong! please re-start the application and try again.'
        return self.xl_data_dictionary